import os
import zipfile
import tempfile
import unittest
from datetime import datetime
from types import SimpleNamespace
from unittest.mock import patch
from xml.etree import ElementTree as ET

import app as app_module
from app import app, round_half_up
from copy1 import copy_excel_with_formatting
from openpyxl import Workbook


class PreviewCalculationTests(unittest.TestCase):
    def setUp(self):
        app.config["TESTING"] = True
        self.client = app.test_client()

    def test_igst_preview_includes_delivery_and_half_up_rounding(self):
        response = self.client.post(
            "/calculate_preview",
            json={
                "items": [{"quantity": 200, "rate": 320}],
                "delivery_charge": 450,
                "tax_type": "IGST",
            },
        )

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["subtotal"], "64000.00")
        self.assertEqual(data["delivery_charge"], "450.00")
        self.assertEqual(data["taxable_amount"], "64450.00")
        self.assertEqual(data["igst_amount"], "3222.50")
        self.assertEqual(data["round_off_value"], "0.50")
        self.assertEqual(data["rounded_total"], "67673.00")
        self.assertEqual(data["amount_in_words"], "Sixty Seven Thousand Six Hundred And Seventy Three Only")

    def test_cgst_sgst_preview_rounds_half_up(self):
        response = self.client.post(
            "/calculate_preview",
            json={
                "items": [{"quantity": 283, "rate": 270}],
                "delivery_charge": 650,
                "tax_type": "CGST_SGST",
            },
        )

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["subtotal"], "76410.00")
        self.assertEqual(data["taxable_amount"], "77060.00")
        self.assertEqual(data["cgst_amount"], "1926.50")
        self.assertEqual(data["sgst_amount"], "1926.50")
        self.assertEqual(data["round_off_value"], "0.00")
        self.assertEqual(data["rounded_total"], "80913.00")

    def test_half_up_rounding_helper(self):
        self.assertEqual(round_half_up(67672.5), 67673)
        self.assertEqual(round_half_up(67672.49), 67672)
        self.assertEqual(round_half_up(67672.51), 67673)


class InvoiceAuditTests(unittest.TestCase):
    def setUp(self):
        app.config["TESTING"] = True
        self.client = app.test_client()
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)

    def test_audit_invoices_reports_duplicates_gaps_and_malformed_files(self):
        filenames = [
            "Invoice_001_2026_27_Alpha.xlsx",
            "Invoice_001_2026_27_Beta.xlsx",
            "Invoice_003_2026_27_Gamma.xlsx",
            "Invoice_bad_name.xlsx",
        ]

        for filename in filenames:
            open(os.path.join(self.tempdir.name, filename), "a").close()

        with patch.object(app_module, "OUTPUT_DIR", self.tempdir.name):
            response = self.client.get("/audit/invoices")

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["total_files"], 4)
        self.assertEqual(data["parsed_files"], 3)
        self.assertEqual(data["malformed_filenames"], ["Invoice_bad_name.xlsx"])
        self.assertEqual(data["duplicate_numbers"], [{"fy": "2026-27", "number": 1, "files": ["Invoice_001_2026_27_Alpha.xlsx", "Invoice_001_2026_27_Beta.xlsx"]}])
        self.assertEqual(data["sequence_gaps"], {"2026-27": [2]})

    def test_next_invoice_number_resets_per_financial_year(self):
        filenames = [
            "Invoice_014_2025_26_Legacy.xlsx",
            "Invoice_078_2025_26_Legacy.xlsx",
            "Invoice_002_2026_27_Current.xlsx",
            "Invoice_009_2026_27_Current.xlsx",
        ]

        for filename in filenames:
            open(os.path.join(self.tempdir.name, filename), "a").close()

        class FixedDate(datetime):
            @classmethod
            def now(cls, tz=None):
                return cls(2026, 4, 8)

        with patch.object(app_module, "OUTPUT_DIR", self.tempdir.name), patch.object(app_module, "datetime", FixedDate):
            response = self.client.get("/api/next_invoice_number")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["next_invoice_number"], "010/2026-27")

    def test_api_invoices_extracts_metadata_from_new_layout_workbook(self):
        cell_values = {
            "E2": "INVOICE No. 005/2026-27",
            "H2": "Date : 08/04/2026",
            "E10": "Mode of Transport: Road",
            "A18": "Aluminium Utensils",
            "F18": 2,
            "C30": "Delivery Charge",
            "E31": "2.50%",
            "E32": "2.50%",
            "I31": 123.45,
            "I32": 123.45,
            "I36": 5000.25,
        }

        class FakeSheet:
            def __getitem__(self, key):
                return SimpleNamespace(value=cell_values.get(key))

        class FakeWorkbook:
            def __init__(self):
                self.active = FakeSheet()
                self.closed = False

            def close(self):
                self.closed = True

        workbook_path = os.path.join(self.tempdir.name, "Invoice_005_2026_27_Widget_Co.xlsx")
        open(workbook_path, "a").close()

        with patch.object(app_module, "OUTPUT_DIR", self.tempdir.name), patch.object(app_module.openpyxl, "load_workbook", return_value=FakeWorkbook()):
            response = self.client.get("/api/invoices")

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(len(data), 1)
        invoice = data[0]
        self.assertEqual(invoice["invoice_number"], "005/2026-27")
        self.assertEqual(invoice["financial_year"], "2026-27")
        self.assertEqual(invoice["buyer_name"], "Widget Co")
        self.assertEqual(invoice["items_count"], 1)
        self.assertEqual(invoice["total_amount"], "5,000.25")
        self.assertEqual(invoice["tax_type"], "CGST+SGST")
        self.assertEqual(invoice["transport_mode"], "Road")

    def test_health_endpoint_reports_ok_when_paths_and_data_are_available(self):
        template_path = os.path.join(self.tempdir.name, "template.xlsx")
        open(template_path, "a").close()
        output_dir = os.path.join(self.tempdir.name, "output")
        pdf_dir = os.path.join(self.tempdir.name, "pdf")
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(pdf_dir, exist_ok=True)

        with patch.object(app_module, "TEMPLATE_EXCEL_FILE", template_path), \
             patch.object(app_module, "OUTPUT_DIR", output_dir), \
             patch.object(app_module, "PDF_OUTPUT_DIR", pdf_dir), \
             patch.object(app_module, "load_data", side_effect=lambda path: []):
            response = self.client.get("/health")

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["status"], "ok")
        self.assertTrue(data["checks"]["template_found"])
        self.assertTrue(data["checks"]["output_dir_writable"])
        self.assertTrue(data["checks"]["pdf_dir_writable"])
        self.assertTrue(data["checks"]["profiles_json_readable"])
        self.assertTrue(data["checks"]["transport_json_readable"])


class InvoiceLayoutTests(unittest.TestCase):
    def test_generated_workbook_stamps_layout_marker(self):
        tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(tempdir.cleanup)

        source_path = os.path.join(tempdir.name, "template.xlsx")
        output_path = os.path.join(tempdir.name, "invoice.xlsx")

        workbook = Workbook()
        sheet = workbook.active
        sheet["C30"] = "Delivery Charge"
        workbook.save(source_path)
        workbook.close()

        copy_excel_with_formatting(
            source_path,
            output_path,
            {
                "buyer_details": ["Buyer : Example"],
                "mode_of_transport": "Road",
                "items": [{"description": "Widget", "quantity": 1, "rate": 100}],
                "delivery_charge": 50,
                "tax_type": "IGST",
                "invoice_number": "001/2026-27",
                "invoice_date": "2026-04-08",
            },
        )

        with zipfile.ZipFile(output_path) as archive:
            sheet_xml = archive.read("xl/worksheets/sheet1.xml")

        root = ET.fromstring(sheet_xml)
        namespaces = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        z1_cell = root.find(".//main:c[@r='Z1']", namespaces)
        self.assertIsNotNone(z1_cell)

        value_node = z1_cell.find(".//main:t", namespaces)
        self.assertIsNotNone(value_node)
        self.assertEqual(value_node.text, "v2")

        columns_node = root.find(".//main:cols", namespaces)
        self.assertIsNotNone(columns_node)
        self.assertIn("hidden=\"1\"", ET.tostring(columns_node, encoding="unicode"))

    def test_layout_detection_prefers_marker_over_old_heuristic(self):
        class FakeCell:
            def __init__(self, value):
                self.value = value

        class FakeSheet:
            def __getitem__(self, key):
                values = {
                    "Z1": "v2",
                    "C30": "Old style summary row",
                }
                return FakeCell(values.get(key))

        self.assertTrue(app_module.is_new_invoice_layout(FakeSheet()))


class InvoiceValidationTests(unittest.TestCase):
    def setUp(self):
        app.config["TESTING"] = True
        app.config["WTF_CSRF_ENABLED"] = False
        self.client = app.test_client()

    def test_generate_invoice_rejects_invalid_item_rows(self):
        profile = {
            "profile_id": "buyer-1",
            "buyer_name": "Example Buyer",
            "buyer_details": ["Buyer : Example"],
            "default_tax_type": "IGST",
        }

        def fake_load_data(path):
            if path == app_module.BUYER_PROFILES_JSON:
                return [profile]
            if path == app_module.TRANSPORT_MODES_JSON:
                return []
            return []

        with patch.object(app_module, "load_data", side_effect=fake_load_data), \
             patch.object(app_module, "TEMPLATE_EXCEL_FILE", "template.xlsx"), \
               patch.object(app_module, "copy_excel_with_formatting") as mocked_copy, \
               patch.object(app_module, "save_new_transport_mode") as mocked_save_transport:
            response = self.client.post(
                "/generate_invoice",
                data={
                    "buyer_profile_id": "buyer-1",
                    "invoice_number": "INV-001",
                    "invoice_date": "2026-04-08",
                    "transport_mode": "Road",
                    "delivery_charge": "0",
                    "item_description[]": "Aluminium Utensils",
                    "item_bags[]": "",
                    "item_quantity[]": "0",
                    "item_rate[]": "100",
                },
                follow_redirects=False,
            )

        self.assertEqual(response.status_code, 302)
        mocked_copy.assert_not_called()
        mocked_save_transport.assert_not_called()

    def test_generate_invoice_rejects_invalid_invoice_number(self):
        profile = {
            "profile_id": "buyer-1",
            "buyer_name": "Example Buyer",
            "buyer_details": ["Buyer : Example"],
            "default_tax_type": "IGST",
        }

        def fake_load_data(path):
            if path == app_module.BUYER_PROFILES_JSON:
                return [profile]
            if path == app_module.TRANSPORT_MODES_JSON:
                return []
            return []

        with patch.object(app_module, "load_data", side_effect=fake_load_data), \
             patch.object(app_module, "TEMPLATE_EXCEL_FILE", "template.xlsx"), \
               patch.object(app_module, "copy_excel_with_formatting") as mocked_copy, \
               patch.object(app_module, "save_new_transport_mode") as mocked_save_transport:
            response = self.client.post(
                "/generate_invoice",
                data={
                    "buyer_profile_id": "buyer-1",
                    "invoice_number": "INV:001",
                    "invoice_date": "2026-04-08",
                    "transport_mode": "Road",
                    "delivery_charge": "0",
                    "item_description[]": "Aluminium Utensils",
                    "item_bags[]": "",
                    "item_quantity[]": "2",
                    "item_rate[]": "100",
                },
                follow_redirects=False,
            )

        self.assertEqual(response.status_code, 302)
        mocked_copy.assert_not_called()
        mocked_save_transport.assert_not_called()

    def test_generate_invoice_rejects_invalid_tax_override(self):
        profile = {
            "profile_id": "buyer-1",
            "buyer_name": "Example Buyer",
            "buyer_details": ["Buyer : Example"],
            "default_tax_type": "IGST",
        }

        def fake_load_data(path):
            if path == app_module.BUYER_PROFILES_JSON:
                return [profile]
            if path == app_module.TRANSPORT_MODES_JSON:
                return []
            return []

        with patch.object(app_module, "load_data", side_effect=fake_load_data), \
             patch.object(app_module, "TEMPLATE_EXCEL_FILE", "template.xlsx"), \
             patch.object(app_module, "copy_excel_with_formatting") as mocked_copy:
            response = self.client.post(
                "/generate_invoice",
                data={
                    "buyer_profile_id": "buyer-1",
                    "invoice_number": "INV001",
                    "invoice_date": "2026-04-08",
                    "transport_mode": "Road",
                    "delivery_charge": "0",
                    "tax_type_override": "BAD_VALUE",
                    "item_description[]": "Aluminium Utensils",
                    "item_bags[]": "",
                    "item_quantity[]": "2",
                    "item_rate[]": "100",
                },
                follow_redirects=False,
            )

        self.assertEqual(response.status_code, 302)
        mocked_copy.assert_not_called()


if __name__ == "__main__":
    unittest.main()
