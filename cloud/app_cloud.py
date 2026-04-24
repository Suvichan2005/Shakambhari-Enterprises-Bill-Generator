"""
Shakambhari Enterprises Invoice Generator - Cloud Version
==========================================================
A Flask-based invoice generation system deployed on Google Cloud with:
- Google Sheets for data storage (buyers, transport modes, invoice records)
- Google Cloud Storage for file storage (Excel, PDF invoices)
- WeasyPrint for PDF generation (no Windows dependency)
"""

import os
import io
import re
import json
import tempfile
import time
import shutil
import subprocess
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict, deque
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, Response, session
from datetime import datetime
import uuid
from num2words import num2words
from typing import Any, List, Dict, Optional
from werkzeug.exceptions import HTTPException

# Cloud integrations
from sheets_db import GoogleSheetsDB, init_sheets_db
from cloud_storage import CloudStorage, init_cloud_storage

# Excel handling
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# PDF generation (cloud-compatible)
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    print("WARNING: WeasyPrint not available. PDF generation will be skipped.")

app = Flask(__name__)

# Secret key for Flask sessions; prefer explicit env var in production.
_secret_key = os.environ.get('FLASK_SECRET_KEY') or os.environ.get('SECRET_KEY')
if not _secret_key:
    _secret_key = os.urandom(32).hex()
    print("WARNING: FLASK_SECRET_KEY is not set. Using ephemeral key; sessions reset on restart.")

app.secret_key = _secret_key
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

secure_cookie_env = os.environ.get('SESSION_COOKIE_SECURE', '').strip().lower()
if secure_cookie_env in {'1', 'true', 'yes'}:
    app.config['SESSION_COOKIE_SECURE'] = True
elif secure_cookie_env in {'0', 'false', 'no'}:
    app.config['SESSION_COOKIE_SECURE'] = False
else:
    # Default to secure only when running in managed cloud runtime.
    app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('K_SERVICE'))

app.config['APP_PASSWORD'] = os.environ.get('APP_PASSWORD', '').strip()
app.config['AUTH_ENABLED'] = bool(app.config['APP_PASSWORD'])

# Basic in-memory per-IP rate limiting for sensitive routes.
RATE_LIMITS = {
    'login': (20, 60),
    'generate_invoice': (30, 60),
    'calculate_preview': (300, 60),
    'api_get_invoice': (180, 60),
    'download_xlsx': (120, 60),
    'download_pdf': (120, 60),
}
_request_windows: Dict[str, deque] = defaultdict(deque)

# Initialize cloud services (lazy loading)
_sheets_db: Optional[GoogleSheetsDB] = None
_cloud_storage: Optional[CloudStorage] = None


AUTH_EXEMPT_ENDPOINTS = {
    'login',
    'logout',
    'health_check',
    'favicon',
    'static',
}


@app.before_request
def require_authentication():
    """Protect routes with a simple session login when APP_PASSWORD is set."""
    if not app.config.get('AUTH_ENABLED'):
        return None

    endpoint = request.endpoint or ''
    if endpoint in AUTH_EXEMPT_ENDPOINTS or endpoint.startswith('static'):
        return None

    if session.get('authenticated'):
        return None

    if request.path.startswith('/api/'):
        return jsonify({'error': 'Authentication required'}), 401

    flash('Please login to continue.', 'warning')
    return redirect(url_for('login', next=request.full_path if request.query_string else request.path))


def _normalize_post_login_target(next_url: str) -> str:
    """Return a safe GET endpoint for post-login navigation."""
    target = (next_url or '').strip()
    if not target.startswith('/'):
        return url_for('index')

    # Never redirect to API or known POST-only endpoints after login.
    if target.startswith('/api/') or target.startswith('/generate_invoice') or target.startswith('/logout'):
        return url_for('index')

    return target


@app.before_request
def enforce_rate_limit():
    """Apply lightweight per-IP throttling to reduce abuse spikes."""
    endpoint = request.endpoint or ''
    if endpoint not in RATE_LIMITS:
        return None

    limit, window_seconds = RATE_LIMITS[endpoint]
    remote = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown').split(',')[0].strip()
    now = time.time()
    key = f"{remote}:{endpoint}"
    bucket = _request_windows[key]

    while bucket and bucket[0] <= now - window_seconds:
        bucket.popleft()

    if len(bucket) >= limit:
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Too many requests. Please retry shortly.'}), 429
        return render_template('error.html',
                              code=429,
                              title='Too Many Requests',
                              message='Please wait a few seconds and try again.'), 429

    bucket.append(now)
    return None


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login route for lightweight app-level protection."""
    if not app.config.get('AUTH_ENABLED'):
        session['authenticated'] = True
        return redirect(url_for('index'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == app.config['APP_PASSWORD']:
            session['authenticated'] = True
            flash('Login successful.', 'success')
            next_url = _normalize_post_login_target(
                request.args.get('next') or request.form.get('next') or url_for('index')
            )
            return redirect(next_url)

        flash('Invalid password.', 'error')

    return render_template('login.html',
                          auth_enabled=app.config.get('AUTH_ENABLED', False),
                          next_url=request.args.get('next', url_for('index')))


@app.route('/logout', methods=['POST'])
def logout():
    """Clear login session."""
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))


def get_sheets_db() -> GoogleSheetsDB:
    """Get or initialize the Google Sheets database connection."""
    global _sheets_db
    if _sheets_db is None:
        _sheets_db = init_sheets_db()
    return _sheets_db


def get_cloud_storage() -> CloudStorage:
    """Get or initialize the Cloud Storage connection."""
    global _cloud_storage
    if _cloud_storage is None:
        _cloud_storage = init_cloud_storage()
    return _cloud_storage


def _safe_filename(filename: str, expected_ext: str) -> str:
    """Normalize and validate download filenames to avoid path manipulation."""
    safe = os.path.basename((filename or '').strip())
    safe = re.sub(r'[^A-Za-z0-9._-]', '_', safe)
    ext = expected_ext.lower()
    if not safe.lower().endswith(ext):
        safe = f"{os.path.splitext(safe)[0]}{ext}"
    return safe


def _filename_from_storage_url(storage_url: str, expected_ext: str) -> str:
    """Extract object name from a gs:// URL and normalize extension."""
    if not storage_url:
        return ''
    fname = storage_url.rsplit('/', 1)[-1]
    return _safe_filename(fname, expected_ext)


def _format_datetime_display(raw_value: str) -> str:
    """Format common timestamp/date strings for UI display."""
    if not raw_value:
        return ''
    text = str(raw_value).strip()
    for fmt in ('%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(text, fmt)
            if fmt == '%Y-%m-%d':
                return dt.strftime('%Y-%m-%d')
            return dt.strftime('%Y-%m-%d %H:%M')
        except ValueError:
            continue
    return text


def _build_invoice_rows(records: List[Dict]) -> List[Dict]:
    """Normalize invoice records for index/dashboard displays."""
    rows: List[Dict] = []
    for rec in records:
        file_url = rec.get('file_url', '')
        pdf_url = rec.get('pdf_url', '')
        xlsx_name = _filename_from_storage_url(file_url, '.xlsx') if file_url else ''
        pdf_name = _filename_from_storage_url(pdf_url, '.pdf') if pdf_url else ''
        items = rec.get('items', [])
        if not isinstance(items, list):
            items = []

        created_raw = rec.get('created_at') or rec.get('invoice_date', '')
        modified_date = _format_datetime_display(created_raw)
        sort_ts = 0
        for fmt in ('%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d'):
            try:
                sort_ts = int(datetime.strptime(str(created_raw), fmt).timestamp())
                break
            except Exception:
                continue

        rows.append({
            **rec,
            'filename': xlsx_name,
            'pdf_filename': pdf_name,
            'items_count': len(items),
            'modified_date': modified_date,
            'sort_ts': sort_ts,
            'transport_mode': extract_transport_core(rec.get('transport_mode', '')),
        })

    rows.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return rows


def _parse_invoice_filename(filename: str) -> Dict[str, str]:
    """Infer invoice number and buyer label from generated invoice filenames."""
    stem = os.path.splitext(filename or '')[0]
    stem = re.sub(r'__\d+$', '', stem)

    if stem.lower().startswith('invoice_'):
        rest = stem[len('invoice_'):]
    else:
        rest = stem

    parts = [p for p in rest.split('_') if p]
    if not parts:
        return {'invoice_number': '', 'buyer_name': ''}

    invoice_number = ''
    buyer_tokens: List[str] = []

    # Invoice_023_2025_26_Prabhat_Aluminium_Industries
    if len(parts) >= 3 and parts[0].isdigit() and re.fullmatch(r'\d{4}', parts[1]) and re.fullmatch(r'\d{2}', parts[2]):
        invoice_number = f"{int(parts[0])}/{parts[1]}-{parts[2]}"
        buyer_tokens = parts[3:]
    # Invoice_8-2026-27_2026-04-17_Buyer_Name or Invoice_1-2025-26_...
    elif re.fullmatch(r'\d+-\d{4}-\d{2}', parts[0] or ''):
        m = re.match(r'^(\d+)-(\d{4})-(\d{2})$', parts[0])
        if m:
            invoice_number = f"{int(m.group(1))}/{m.group(2)}-{m.group(3)}"
        buyer_tokens = parts[1:]
    # Invoice_001_2026_27_Tirupati_Udyog (already handled above) or fallback first token only
    elif parts[0].isdigit():
        invoice_number = str(int(parts[0]))
        buyer_tokens = parts[1:]
    else:
        buyer_tokens = parts

    # Skip date token if present after invoice token.
    if buyer_tokens and re.fullmatch(r'\d{4}-\d{2}-\d{2}', buyer_tokens[0]):
        buyer_tokens = buyer_tokens[1:]

    buyer_name = ' '.join(buyer_tokens).replace('-', ' ').strip()
    buyer_name = re.sub(r'\s+', ' ', buyer_name)

    return {
        'invoice_number': invoice_number,
        'buyer_name': buyer_name,
    }


def _extract_invoice_data_from_xlsx_bytes(file_bytes: bytes, filename: str = '') -> Dict[str, Any]:
    """Extract invoice payload from an XLSX file stored in GCS."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    try:
        sheet = wb.active

        invoice_num_raw = str(sheet['E2'].value or '').strip()
        invoice_number = re.sub(r'(?i)^\s*invoice\s*no\.?\s*', '', invoice_num_raw).strip(' :-')

        invoice_date_raw = str(sheet['H2'].value or '').strip()
        invoice_date_text = re.sub(r'(?i)^\s*date\s*', '', invoice_date_raw).strip(' :-')
        invoice_date = ''
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(invoice_date_text, fmt)
                invoice_date = dt.strftime('%Y-%m-%d')
                break
            except ValueError:
                continue

        buyer_details: List[str] = []
        for row in range(8, 16):
            value = sheet[f'A{row}'].value
            if value not in (None, ''):
                buyer_details.append(str(value).strip())

        buyer_name = ''
        for line in buyer_details:
            norm = line.strip().lower()
            if norm in {'buyer :', 'buyer:'}:
                continue
            buyer_name = line.strip()
            break

        buyer_gstin = ''
        for line in buyer_details:
            m = re.search(r'GSTIN\s*[-:]\s*([A-Z0-9]+)', str(line), re.IGNORECASE)
            if m:
                buyer_gstin = m.group(1).upper()
                break

        transport_mode = extract_transport_core(str(sheet['E10'].value or '').strip())

        items: List[Dict[str, Any]] = []
        for row in range(18, 28):
            desc_raw = str(sheet[f'A{row}'].value or '').strip()
            hsn_val = sheet[f'H{row}'].value or sheet[f'B{row}'].value
            hsn = str(hsn_val).strip() if hsn_val not in (None, '') else ''
            qty = safe_float(sheet[f'F{row}'].value, 0.0)
            rate = safe_float(sheet[f'G{row}'].value, 0.0)

            if desc_raw or qty or rate:
                base_desc = re.sub(r'^\d+\.\s*', '', desc_raw).strip()
                bags = ''
                bag_match = re.search(r'\((\d+(?:\.\d+)?)\s*Bags?\)', base_desc, re.IGNORECASE)
                if bag_match:
                    bags = bag_match.group(1)
                    base_desc = re.sub(r'\s*\(\d+(?:\.\d+)?\s*Bags?\)', '', base_desc, flags=re.IGNORECASE).strip()

                items.append({
                    'description': base_desc,
                    'bags': bags,
                    'hsn': hsn,
                    'quantity': qty,
                    'rate': rate,
                })

        delivery_charge = 0.0
        label_30 = str(sheet['C30'].value or '').strip().lower()
        if 'delivery' in label_30:
            delivery_charge = safe_float(sheet['I30'].value, 0.0)

        tax_type = 'IGST'
        rate_31 = str(sheet['E31'].value or '').strip()
        rate_32 = str(sheet['E32'].value or '').strip()
        if rate_31.startswith('2.50') or rate_32.startswith('2.50'):
            tax_type = 'CGST_SGST'

        parsed_from_name = _parse_invoice_filename(filename)
        if not invoice_number:
            invoice_number = parsed_from_name.get('invoice_number', '')
        if not buyer_name:
            buyer_name = parsed_from_name.get('buyer_name', '')

        return {
            'invoice_number': invoice_number,
            'invoice_date': invoice_date,
            'buyer_name': buyer_name,
            'buyer_gstin': buyer_gstin,
            'buyer_details': buyer_details,
            'transport_mode': transport_mode,
            'delivery_charge': delivery_charge,
            'items': items,
            'tax_type': tax_type,
            'filename': filename,
            'pdf_filename': filename.replace('.xlsx', '.pdf') if filename.lower().endswith('.xlsx') else '',
        }
    finally:
        wb.close()


def _merge_with_storage_rows(sheet_rows: List[Dict], storage_rows: List[Dict]) -> List[Dict]:
    """Merge invoice metadata with bucket files so the modal shows all available files."""
    merged = list(sheet_rows)
    existing = {row.get('filename', '') for row in sheet_rows if row.get('filename')}

    for blob in storage_rows:
        filename = blob.get('filename', '')
        if not filename or filename in existing:
            continue

        parsed = _parse_invoice_filename(filename)
        updated = blob.get('updated')
        modified_date = ''
        if updated:
            try:
                modified_date = updated.strftime('%Y-%m-%d %H:%M')
            except Exception:
                modified_date = str(updated)

        merged.append({
            'filename': filename,
            'pdf_filename': filename.replace('.xlsx', '.pdf'),
            'invoice_number': parsed.get('invoice_number', ''),
            'buyer_name': parsed.get('buyer_name', ''),
            'modified_date': modified_date,
            'sort_ts': int(updated.timestamp()) if updated else 0,
            'items_count': 0,
            'tax_type': '',
            'transport_mode': '',
            'total_amount': '',
        })

    merged.sort(key=lambda x: x.get('sort_ts', 0), reverse=True)
    return merged


# ===================== UTILITY FUNCTIONS =====================

def _financial_year_suffix(today: datetime = None) -> str:
    """Get the financial year suffix like /2025-26."""
    today = today or datetime.now()
    year = today.year
    if today.month >= 4:
        start = year
        end = year + 1
    else:
        start = year - 1
        end = year
    return f"/{start}-{str(end)[-2:]}"


def suggest_next_invoice_number() -> str:
    """Suggest the next invoice number based on the last one."""
    db = get_sheets_db()
    last_num = db.get_last_invoice_number()
    
    if not last_num:
        fy = _financial_year_suffix()
        return f"1{fy}"
    
    # Extract numeric part
    match = re.match(r'^(\d+)', last_num)
    if match:
        num = int(match.group(1)) + 1
        fy = _financial_year_suffix()
        return f"{num}{fy}"
    
    return last_num


def format_date_for_invoice(date_str: str) -> str:
    """Convert YYYY-MM-DD to DD/MM/YYYY for invoice display."""
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.strftime('%d/%m/%Y')
    except ValueError:
        return date_str


def amount_in_words(amount: float) -> str:
    """Convert amount to words for invoice."""
    try:
        rupees = int(amount)
        paise = int(round((amount - rupees) * 100))
        
        if rupees == 0 and paise == 0:
            return "Zero Only"
        
        words = num2words(rupees, lang='en_IN').title()
        words = words.replace(',', '')
        
        if paise > 0:
            paise_words = num2words(paise, lang='en_IN').title()
            return f"Rupees {words} and {paise_words} Paise Only"
        
        return f"Rupees {words} Only"
    except Exception:
        return f"Rupees {amount} Only"


def safe_float(value: Any, default: float = 0.0) -> float:
    """Safely convert a value to float with a fallback."""
    try:
        if value is None or value == '':
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def round_half_up(value: float) -> int:
    """Round to nearest integer with .5 always rounding up (local app parity)."""
    return int(Decimal(str(value)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def calculate_invoice_totals(items: List[Dict], tax_type: str = 'IGST', delivery_charge: float = 0.0) -> Dict:
    """Calculate invoice totals from items."""
    subtotal = sum(safe_float(item.get('quantity', 0), 0.0) * safe_float(item.get('rate', 0), 0.0) for item in items)
    delivery_charge = max(0.0, safe_float(delivery_charge, 0.0))
    taxable_amount = subtotal + delivery_charge
    
    if tax_type == 'IGST':
        igst = round(taxable_amount * 0.05, 2)
        tax_amount = igst
        cgst = sgst = 0
    else:
        igst = 0
        cgst = round(taxable_amount * 0.025, 2)
        sgst = round(taxable_amount * 0.025, 2)
        tax_amount = cgst + sgst
    
    total_before_round = round(taxable_amount + tax_amount, 2)
    rounded_total = round(total_before_round)
    round_off = round(rounded_total - total_before_round, 2)
    
    return {
        'subtotal': round(subtotal, 2),
        'delivery_charge': round(delivery_charge, 2),
        'taxable_amount': round(taxable_amount, 2),
        'igst_amount': round(igst, 2),
        'cgst_amount': round(cgst, 2),
        'sgst_amount': round(sgst, 2),
        'tax_amount': round(tax_amount, 2),
        'round_off_value': round(round_off, 2),
        'rounded_total': rounded_total,
        'amount_in_words': amount_in_words(rounded_total)
    }


def extract_transport_core(mode: str) -> str:
    """Extract core transport mode without prefix."""
    if not mode:
        return ''

    value = mode.strip()
    variants = [
        'mode of transport:',
        'mode of transport :',
        'mode of transports:',
        'mode of transports :',
        'transport:',
        'transport :',
    ]

    lowered = value.lower()
    for prefix in variants:
        if lowered.startswith(prefix):
            value = value[len(prefix):].strip(' -:')
            break

    return value.strip()


def normalize_transport_mode(mode: str) -> str:
    """Return canonical transport string used in Excel and storage."""
    core = extract_transport_core(mode)
    if not core:
        return ''
    return f"Mode of Transport: {core}"


def _normalize_name(value: str) -> str:
    """Normalize names for robust profile matching."""
    if not value:
        return ''
    return re.sub(r'\s+', ' ', value.strip().lower())


def _match_buyer_profile(invoice: Dict, buyers: List[Dict]) -> Optional[Dict]:
    """Find the best buyer profile match for a stored invoice record."""
    if not buyers:
        return None

    profile_id = (invoice.get('buyer_profile_id') or '').strip()
    if profile_id:
        for buyer in buyers:
            if (buyer.get('profile_id') or '').strip() == profile_id:
                return buyer

    target_gstin = (invoice.get('buyer_gstin') or '').strip().upper()
    if target_gstin:
        for buyer in buyers:
            if (buyer.get('gstin') or '').strip().upper() == target_gstin:
                return buyer

    target_name = _normalize_name(invoice.get('buyer_name', ''))
    if target_name:
        for buyer in buyers:
            buyer_name = _normalize_name(buyer.get('buyer_name', ''))
            if buyer_name and (buyer_name == target_name or buyer_name in target_name or target_name in buyer_name):
                return buyer

    details = invoice.get('buyer_details', [])
    if isinstance(details, list):
        for line in details:
            line_norm = _normalize_name(str(line))
            if not line_norm or line_norm in {'buyer:', 'buyer :'}:
                continue
            for buyer in buyers:
                buyer_name = _normalize_name(buyer.get('buyer_name', ''))
                if buyer_name and (buyer_name == line_norm or buyer_name in line_norm or line_norm in buyer_name):
                    return buyer

    return None


# ===================== EXCEL GENERATION =====================

def generate_invoice_excel(invoice_data: Dict) -> bytes:
    """
    Generate an Excel invoice from the template stored in Cloud Storage.
    Returns the Excel file as bytes.
    """
    storage = get_cloud_storage()
    
    # Download template
    template_result = storage.download_template()
    if not template_result:
        raise Exception("Invoice template not found in Cloud Storage")
    
    template_bytes, template_name = template_result
    
    source_wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    source_sheet = source_wb.active

    dest_wb = openpyxl.Workbook()
    if dest_wb.sheetnames:
        dest_wb.remove(dest_wb.active)
    dest_sheet = dest_wb.create_sheet(title=source_sheet.title)

    # Copy page setup and all cells/styles exactly (local parity behavior).
    dest_sheet.page_setup = copy(source_sheet.page_setup)
    dest_sheet.page_margins = copy(source_sheet.page_margins)

    for row in source_sheet.iter_rows():
        for source_cell in row:
            dest_cell = dest_sheet.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
            if source_cell.has_style:
                dest_cell.font = copy(source_cell.font)
                dest_cell.border = copy(source_cell.border)
                dest_cell.fill = copy(source_cell.fill)
                dest_cell.number_format = source_cell.number_format
                dest_cell.protection = copy(source_cell.protection)
                dest_cell.alignment = copy(source_cell.alignment)

    # Clear residual thin line on Salt lake Sector 2 (Column B, Row 41)
    if dest_sheet['B41'].border:
        new_border = copy(dest_sheet['B41'].border)
        new_border.left = Side(style=None)
        dest_sheet['B41'].border = new_border

    dest_sheet['E2'] = f"INVOICE No. {invoice_data['invoice_number']}"
    dest_sheet['H2'] = f"Date : {invoice_data['invoice_date_display']}"

    buyer_details = invoice_data.get('buyer_details', [])
    for i, detail in enumerate(buyer_details[:8]):
        dest_sheet[f'A{8+i}'] = detail

    # Clear neighbor cells first so no legacy template text leaks into output.
    for cell_ref in ('F10', 'G10', 'H10'):
        dest_sheet[cell_ref] = ''
    dest_sheet['E10'] = normalize_transport_mode(invoice_data.get('transport_mode', ''))

    items = invoice_data.get('items', [])
    first_item_row = 18
    template_hsn = source_sheet[f'H{first_item_row}'].value or source_sheet[f'B{first_item_row}'].value
    item_rows: List[int] = []

    for idx, item in enumerate(items[:10]):
        row_num = first_item_row + idx
        item_rows.append(row_num)

        description = item.get('description', '')
        if len(items) > 1 and description and not description[0].isdigit():
            description = f"{idx + 1}. {description}"

        hsn_value = item.get('hsn') or template_hsn
        quantity = safe_float(item.get('quantity', 0), 0.0)
        rate = safe_float(item.get('rate', 0), 0.0)

        dest_sheet[f'A{row_num}'] = description
        if hsn_value is not None:
            dest_sheet[f'H{row_num}'] = hsn_value
        dest_sheet[f'F{row_num}'] = quantity
        dest_sheet[f'F{row_num}'].number_format = '0.000'
        dest_sheet[f'G{row_num}'] = rate
        dest_sheet[f'G{row_num}'].number_format = '0.00'
        dest_sheet[f'I{row_num}'] = f'=F{row_num}*G{row_num}'
        dest_sheet[f'I{row_num}'].number_format = '0.00'

        # Explicitly copy all styles from the first item row (row 18) to ensure 
        # missing borders, fonts (like bold HSN), and alignments apply to rows 2+.
        if row_num > first_item_row:
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                src_cell = dest_sheet[f'{col_letter}{first_item_row}']
                tgt_cell = dest_sheet[f'{col_letter}{row_num}']
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    
                    # Copy border but strip the top border so we don't draw lines between items
                    new_border = copy(src_cell.border)
                    new_border.top = openpyxl.styles.borders.Side(style=None)
                    tgt_cell.border = new_border
                    
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.alignment = copy(src_cell.alignment)
                    if not tgt_cell.number_format or tgt_cell.number_format == 'General':
                        tgt_cell.number_format = src_cell.number_format

    if len(item_rows) == 1:
        subtotal_formula = f'=I{item_rows[0]}'
    elif len(item_rows) > 1:
        subtotal_formula = f'=SUM(I{item_rows[0]}:I{item_rows[-1]})'
    else:
        subtotal_formula = '=0'

    dest_sheet['Z1'] = 'v2'
    dest_sheet.column_dimensions['Z'].hidden = True

    dest_sheet['I29'] = subtotal_formula
    dest_sheet['I29'].number_format = '0.00'

    delivery_charge = max(0.0, safe_float(invoice_data.get('delivery_charge', 0), 0.0))
    dest_sheet['C30'] = 'Delivery Charge'
    dest_sheet['E30'] = ''
    dest_sheet['I30'] = delivery_charge
    dest_sheet['I30'].number_format = '0.00'

    tax_type = invoice_data.get('tax_type', 'IGST')
    tax_base_formula = '(I29+I30)'
    if tax_type == 'IGST':
        dest_sheet['C31'] = 'G.S.T SALES I.G.S.T @'
        dest_sheet['E31'] = '5.00%'
        dest_sheet['I31'] = f'=ROUND({tax_base_formula}*0.05, 2)'
        dest_sheet['I31'].number_format = '0.00'

        dest_sheet['C32'] = 'G.S.T SALES C.G.S.T @'
        dest_sheet['E32'] = '0.00%'
        dest_sheet['I32'] = 0.0
        dest_sheet['I32'].number_format = '0.00'

        dest_sheet['C33'] = ''
        dest_sheet['E33'] = ''
        dest_sheet['I33'] = ''
    else:
        dest_sheet['C31'] = 'G.S.T SALES C.G.S.T @'
        dest_sheet['E31'] = '2.50%'
        dest_sheet['I31'] = f'=ROUND({tax_base_formula}*0.025, 2)'
        dest_sheet['I31'].number_format = '0.00'

        dest_sheet['C32'] = 'G.S.T SALES S.G.S.T @'
        dest_sheet['E32'] = '2.50%'
        dest_sheet['I32'] = f'=ROUND({tax_base_formula}*0.025, 2)'
        dest_sheet['I32'].number_format = '0.00'

        dest_sheet['C33'] = ''
        dest_sheet['E33'] = ''
        dest_sheet['I33'] = ''

    dest_sheet['I38'] = '=I29+I30+I31+I32'
    dest_sheet.row_dimensions[38].hidden = True
    dest_sheet['I34'] = '=ROUND(I38,0)-I38'
    dest_sheet['I34'].number_format = '0.00'
    dest_sheet['I35'] = '=ROUND(I38,0)'
    dest_sheet['I35'].number_format = '0.00'
    dest_sheet['I36'] = ''

    subtotal = sum(safe_float(item.get('quantity', 0), 0.0) * safe_float(item.get('rate', 0), 0.0) for item in items)
    tax_base_value = subtotal + delivery_charge
    if tax_type == 'IGST':
        igst_value = tax_base_value * 0.05
        dest_sheet.row_dimensions[31].hidden = igst_value <= 0
        dest_sheet.row_dimensions[32].hidden = True
    else:
        cgst_value = tax_base_value * 0.025
        sgst_value = tax_base_value * 0.025
        dest_sheet.row_dimensions[31].hidden = cgst_value <= 0
        dest_sheet.row_dimensions[32].hidden = sgst_value <= 0
    dest_sheet.row_dimensions[30].hidden = delivery_charge <= 0
    dest_sheet.row_dimensions[33].hidden = False

    if tax_type == 'IGST':
        tax_amount = tax_base_value * 0.05
    else:
        tax_amount = tax_base_value * 0.05
    total_before_round = subtotal + delivery_charge + tax_amount
    rounded_total = round_half_up(total_before_round)
    # Remove any old amount words text from adjacent cells.
    for col in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'):
        dest_sheet[f'{col}37'] = ''

    if rounded_total > 0:
        words = num2words(int(rounded_total), lang='en_IN').replace('-', ' ').replace(',', ' ').title()
        amount_words = f"AMOUNT : {words} Only"
    else:
        amount_words = 'AMOUNT : Zero Only'
    dest_sheet['A37'] = amount_words

    for col_letter, source_dim in source_sheet.column_dimensions.items():
        dest_dim = dest_sheet.column_dimensions[col_letter]
        dest_dim.width = source_dim.width
        dest_dim.hidden = source_dim.hidden
        dest_dim.outline_level = source_dim.outline_level
        dest_dim.collapsed = source_dim.collapsed

    for row_idx, source_dim in source_sheet.row_dimensions.items():
        if row_idx in {30, 31, 32, 33, 38}:
            continue
        dest_dim = dest_sheet.row_dimensions[row_idx]
        dest_dim.height = source_dim.height
        dest_dim.hidden = source_dim.hidden
        dest_dim.outline_level = source_dim.outline_level
        dest_dim.collapsed = source_dim.collapsed

    for merged_cell_range in source_sheet.merged_cells.ranges:
        dest_sheet.merge_cells(str(merged_cell_range))

    output = io.BytesIO()
    dest_wb.save(output)
    output.seek(0)
    source_wb.close()
    dest_wb.close()
    return output.read()


def generate_invoice_pdf(invoice_data: Dict) -> Optional[bytes]:
    """
    Generate a PDF invoice using WeasyPrint.
    Returns the PDF file as bytes.
    """
    if not WEASYPRINT_AVAILABLE:
        return None
    
    try:
        # Render HTML template
        html_content = render_template('invoice_pdf_template.html', invoice=invoice_data)

        # Convert to PDF
        pdf_bytes = HTML(string=html_content).write_pdf()
        return pdf_bytes
    except Exception as exc:
        app.logger.exception('PDF generation failed: %s', exc)
        return None


def generate_pdf_from_excel(excel_bytes: bytes, excel_filename: str) -> Optional[bytes]:
    """
    Generate PDF directly from XLSX using LibreOffice headless.
    This provides visual output closest to the spreadsheet layout.
    """
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if not soffice:
        app.logger.error('LibreOffice (soffice) is not installed in this runtime.')
        return None

    safe_xlsx_name = _safe_filename(excel_filename or 'invoice.xlsx', '.xlsx')

    with tempfile.TemporaryDirectory(prefix='xlsx_to_pdf_') as tmpdir:
        xlsx_path = os.path.join(tmpdir, safe_xlsx_name)
        with open(xlsx_path, 'wb') as f:
            f.write(excel_bytes)

        cmd = [
            soffice,
            '--headless',
            '--nologo',
            '--nolockcheck',
            '--nodefault',
            '--nofirststartwizard',
            '--convert-to', 'pdf:calc_pdf_Export',
            '--outdir', tmpdir,
            xlsx_path,
        ]

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=120,
                check=False,
            )
        except Exception as exc:
            app.logger.exception('Failed to run LibreOffice conversion: %s', exc)
            return None

        if result.returncode != 0:
            app.logger.error(
                'LibreOffice conversion failed (code=%s): stdout=%s stderr=%s',
                result.returncode,
                (result.stdout or '').strip(),
                (result.stderr or '').strip(),
            )
            return None

        expected_pdf = os.path.splitext(xlsx_path)[0] + '.pdf'
        pdf_path = expected_pdf if os.path.exists(expected_pdf) else ''
        if not pdf_path:
            # Fallback in case LibreOffice produces a different basename.
            for name in os.listdir(tmpdir):
                if name.lower().endswith('.pdf'):
                    pdf_path = os.path.join(tmpdir, name)
                    break

        if not pdf_path or not os.path.exists(pdf_path):
            app.logger.error('LibreOffice conversion completed but no PDF output was found.')
            return None

        with open(pdf_path, 'rb') as f:
            return f.read()


# ===================== ROUTE HANDLERS =====================

@app.route('/')
def index():
    """Main invoice generation page."""
    db = get_sheets_db()
    
    buyer_profiles = db.get_all_buyers()
    transport_modes = db.get_all_transport_modes()
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    # Sort buyers by name
    buyer_profiles.sort(key=lambda p: p.get('buyer_name', '').lower())
    
    # Normalize transport modes and enrich from invoice history for better suggestions.
    invoice_transport_modes = [inv.get('transport_mode', '') for inv in db.get_all_invoices(limit=500)]
    combined_transport_modes = [m for m in (transport_modes + invoice_transport_modes) if m]
    transport_cores = list(set(extract_transport_core(m) for m in combined_transport_modes if extract_transport_core(m)))
    transport_cores.sort()
    
    suggestion = suggest_next_invoice_number()
    recent_invoices = _build_invoice_rows(db.get_all_invoices(limit=2000))
    try:
        storage_rows = get_cloud_storage().list_invoices(limit=2000)
        recent_invoices = _merge_with_storage_rows(recent_invoices, storage_rows)
    except Exception as exc:
        app.logger.warning('Could not list storage invoices for modal merge: %s', exc)

    bucket = os.environ.get('GCS_BUCKET_NAME', '')
    project = os.environ.get('GOOGLE_CLOUD_PROJECT', '')
    bucket_base = f"https://console.cloud.google.com/storage/browser/{bucket}"
    project_suffix = f"?project={project}" if project else ''
    
    # Check if loading a specific invoice
    load_invoice_number = request.args.get('load', '')
    preload_invoice = None
    if load_invoice_number:
        preload_invoice = db.get_invoice(load_invoice_number)
    
    return render_template('index.html',
                          buyer_profiles=buyer_profiles,
                          transport_modes=transport_cores,
                          today_date=today_date,
                          suggested_invoice_number=suggestion,
                          recent_invoices=recent_invoices,
                          preload_invoice=preload_invoice,
                          open_records=(request.args.get('open_records') == '1'),
                          bucket_console_url=f"{bucket_base}{project_suffix}" if bucket else '',
                          invoices_folder_url=f"{bucket_base}/invoices/{project_suffix}" if bucket else '',
                          pdfs_folder_url=f"{bucket_base}/pdfs/{project_suffix}" if bucket else '')


@app.route('/dashboard')
def dashboard():
    """Dashboard is merged into index modal; keep this route as a compatibility redirect."""
    return redirect(url_for('index', open_records='1'))


@app.route('/generate_invoice', methods=['POST'])
def generate_invoice():
    """Generate an invoice from form data."""
    try:
        db = get_sheets_db()
        storage = get_cloud_storage()
        
        # Get form data
        buyer_profile_id = request.form.get('buyer_profile_id')
        if not buyer_profile_id:
            flash("Please select a buyer profile.", "error")
            return redirect(url_for('index'))
        
        invoice_number = request.form.get('invoice_number', '').strip()
        invoice_date_str = request.form.get('invoice_date', '')
        transport_mode_input = request.form.get('transport_mode_input', request.form.get('transport_mode', '')).strip()
        transport_mode = normalize_transport_mode(transport_mode_input)
        delivery_charge = safe_float(request.form.get('delivery_charge', '0').strip(), 0.0)
        if delivery_charge < 0:
            flash("Delivery charge cannot be negative.", "error")
            return redirect(url_for('index'))
        tax_type_override = request.form.get('tax_type_override', 'PROFILE_DEFAULT')
        
        # Get buyer profile
        buyer = db.get_buyer(buyer_profile_id)
        if not buyer:
            flash("Buyer profile not found.", "error")
            return redirect(url_for('index'))
        
        # Determine tax type
        if tax_type_override == 'PROFILE_DEFAULT':
            tax_type = buyer.get('default_tax_type', 'IGST')
        else:
            tax_type = tax_type_override
        
        # Parse items
        descriptions = request.form.getlist('item_description[]')
        bags = request.form.getlist('item_bags[]')
        item_hsns = request.form.getlist('item_hsn[]')
        quantities = request.form.getlist('item_quantity[]')
        rates = request.form.getlist('item_rate[]')
        
        items = []
        for i in range(len(descriptions)):
            try:
                qty = float(quantities[i]) if quantities[i] else 0
                rate = float(rates[i]) if rates[i] else 0
                
                if qty > 0 or rate > 0:
                    desc = descriptions[i].strip()
                    bag_val = bags[i].strip() if i < len(bags) and bags[i] else ''
                    
                    # Prevent duplication by stripping existing bags suffix
                    desc = re.sub(r'\s*\(\s*\d+(?:\.\d+)?\s*Bags?\s*\)', '', desc, flags=re.IGNORECASE).strip()
                    
                    if bag_val:
                        desc += f" ({bag_val} Bags)"
                    
                    hsn = item_hsns[i].strip() if i < len(item_hsns) else ''
                    items.append({
                        'description': desc,
                        'bags': bag_val,
                        'hsn': hsn,
                        'quantity': qty,
                        'rate': rate,
                        'amount': qty * rate
                    })
            except (ValueError, IndexError):
                continue
        
        if not items:
            flash("Please add at least one item.", "error")
            return redirect(url_for('index'))
        
        # Calculate totals
        totals = calculate_invoice_totals(items, tax_type, delivery_charge=delivery_charge)
        
        # Prepare invoice data
        invoice_data = {
            'invoice_number': invoice_number,
            'invoice_date': invoice_date_str,
            'invoice_date_display': format_date_for_invoice(invoice_date_str),
            'buyer_name': buyer['buyer_name'],
            'buyer_gstin': buyer.get('gstin', ''),
            'buyer_details': buyer.get('buyer_details', []),
            'items': items,
            'transport_mode': transport_mode,
            'delivery_charge': delivery_charge,
            'tax_type': tax_type,
            **totals
        }
        
        # Generate Excel file
        excel_bytes = generate_invoice_excel(invoice_data)
        
        # Generate filename
        safe_buyer = ''.join(c if c.isalnum() else '_' for c in buyer['buyer_name'][:20])
        filename = f"Invoice_{invoice_number.replace('/', '-')}_{invoice_date_str}_{safe_buyer}.xlsx"
        
        # Upload to Cloud Storage
        xlsx_url = storage.upload_invoice_xlsx(excel_bytes, filename)
        
        # Generate and upload PDF
        pdf_url = ''
        pdf_filename = filename.replace('.xlsx', '.pdf')
        
        pdf_bytes = generate_pdf_from_excel(excel_bytes, filename)
        if pdf_bytes:
            pdf_url = storage.upload_invoice_pdf(pdf_bytes, pdf_filename)
        else:
            app.logger.error('Skipping PDF upload because XLSX-to-PDF conversion failed for %s', filename)
        
        # Save invoice record to Google Sheets
        invoice_record = {
            'invoice_number': invoice_number,
            'invoice_date': invoice_date_str,
            'buyer_name': buyer['buyer_name'],
            'buyer_gstin': buyer.get('gstin', ''),
            'items': items,
            'subtotal': totals['subtotal'],
            'tax_type': tax_type,
            'tax_amount': totals['tax_amount'],
            'total_amount': totals['rounded_total'],
            'transport_mode': transport_mode,
            'file_url': xlsx_url,
            'pdf_url': pdf_url
        }
        db.save_invoice(invoice_record)
        
        # Save new transport mode if provided
        if transport_mode:
            db.add_transport_mode(transport_mode)
        
        flash(f"Invoice {invoice_number} generated successfully!", "success")

        excel_download_url = url_for('download_xlsx', filename=filename)
        pdf_download_url = url_for('download_pdf', filename=pdf_filename) if pdf_url else ''
        
        return render_template('success.html',
                              filename=filename,
                              invoice_number=invoice_number,
                      excel_url=excel_download_url,
                      pdf_url=pdf_download_url,
                      is_pdf=bool(pdf_url))
        
    except Exception as e:
        flash(f"Error generating invoice: {str(e)}", "error")
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))


@app.route('/download/xlsx/<filename>')
def download_xlsx(filename):
    """Download an Excel invoice."""
    safe_name = _safe_filename(filename, '.xlsx')
    storage = get_cloud_storage()
    file_bytes = storage.download_invoice_xlsx(safe_name)
    
    if not file_bytes:
        flash("File not found.", "error")
        return redirect(url_for('index'))
    
    return Response(
        file_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={safe_name}'}
    )


@app.route('/download/pdf/<filename>')
def download_pdf(filename):
    """Download a PDF invoice."""
    safe_name = _safe_filename(filename, '.pdf')
    storage = get_cloud_storage()
    file_bytes = storage.download_invoice_pdf(safe_name)
    
    if not file_bytes:
        flash("PDF not found.", "error")
        return redirect(url_for('index'))
    
    return Response(
        file_bytes,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={safe_name}'}
    )


# ===================== BUYER PROFILE ROUTES =====================

@app.route('/profiles')
def list_profiles():
    """List all buyer profiles."""
    db = get_sheets_db()
    profiles = db.get_all_buyers()
    profiles.sort(key=lambda p: p.get('buyer_name', '').lower())
    return render_template('list_profiles.html', profiles=profiles)


@app.route('/profile', methods=['GET', 'POST'])
@app.route('/profile/<profile_id>', methods=['GET', 'POST'])
def manage_profile(profile_id=None):
    """Create or edit a buyer profile."""
    db = get_sheets_db()
    
    is_new_profile = profile_id is None
    profile = None
    
    if profile_id:
        profile = db.get_buyer(profile_id)
        if not profile:
            flash("Profile not found.", "error")
            return redirect(url_for('list_profiles'))
    
    if request.method == 'POST':
        buyer_name = request.form.get('buyer_name', '').strip()
        buyer_details_str = request.form.get('buyer_details_textarea', '')
        buyer_details = [line.strip() for line in buyer_details_str.split('\n') if line.strip()]
        gstin = request.form.get('gstin', '').strip().upper()
        default_tax_type = request.form.get('default_tax_type', 'IGST')
        
        if not buyer_name:
            flash("Buyer Name is required.", "error")
            profile_data = {
                'buyer_name': buyer_name,
                'buyer_details_textarea': buyer_details_str,
                'buyer_details': buyer_details,
                'gstin': gstin,
                'default_tax_type': default_tax_type,
                'profile_id': profile_id or ''
            }
            return render_template('profile_form.html', profile=profile_data,
                                 is_new_profile=is_new_profile)
        
        if is_new_profile:
            # Generate profile ID
            new_profile_id = gstin if gstin else f"{buyer_name.replace(' ', '_')}_{uuid.uuid4().hex[:8]}"
        else:
            new_profile_id = profile_id
        
        profile_data = {
            'profile_id': new_profile_id,
            'buyer_name': buyer_name,
            'buyer_details': buyer_details,
            'gstin': gstin,
            'default_tax_type': default_tax_type
        }
        
        if db.save_buyer(profile_data):
            flash(f"Profile '{buyer_name}' saved successfully!", "success")
            return redirect(url_for('list_profiles'))
        else:
            flash("Error saving profile.", "error")
    
    # GET request
    if profile:
        profile['buyer_details_textarea'] = '\n'.join(profile.get('buyer_details', []))
    else:
        profile = {
            'buyer_name': '',
            'buyer_details': [],
            'buyer_details_textarea': '',
            'gstin': '',
            'default_tax_type': 'IGST',
            'profile_id': ''
        }
    
    return render_template('profile_form.html', profile=profile, is_new_profile=is_new_profile)


@app.route('/profile/<profile_id>/delete', methods=['POST'])
def delete_profile(profile_id):
    """Delete a buyer profile."""
    db = get_sheets_db()
    
    if db.delete_buyer(profile_id):
        flash("Profile deleted successfully.", "success")
    else:
        flash("Profile not found.", "error")
    
    return redirect(url_for('list_profiles'))


# ===================== API ROUTES =====================

@app.route('/api/calculate', methods=['POST'])
def calculate_preview():
    """API endpoint to calculate invoice preview."""
    data = request.get_json() or {}
    items = data.get('items', [])
    tax_type = data.get('tax_type', 'IGST')
    delivery_charge = safe_float(data.get('delivery_charge', 0), 0.0)
    
    totals = calculate_invoice_totals(items, tax_type, delivery_charge=delivery_charge)
    return jsonify(totals)


@app.route('/calculate_preview', methods=['POST'])
def calculate_preview_route():
    """Backward-compatible alias used by older templates."""
    return calculate_preview()


@app.route('/api/invoice/<path:invoice_number>')
def api_get_invoice(invoice_number):
    """API endpoint to get invoice data for loading."""
    db = get_sheets_db()
    invoice = db.get_invoice(invoice_number)
    
    if not invoice:
        return jsonify({'error': 'Invoice not found'}), 404

    invoice['filename'] = _filename_from_storage_url(invoice.get('file_url', ''), '.xlsx')
    invoice['pdf_filename'] = _filename_from_storage_url(invoice.get('pdf_url', ''), '.pdf')
    invoice['transport_mode'] = extract_transport_core(invoice.get('transport_mode', ''))

    tax_amt = float(invoice.get('tax_amount', 0) or 0)
    subt = float(invoice.get('subtotal', 0) or 0)
    if tax_amt > 0:
        est_dc = (tax_amt / 0.05) - subt
        if abs(est_dc - round(est_dc)) < 0.2:
            invoice['delivery_charge'] = float(round(est_dc))
        else:
            invoice['delivery_charge'] = round(est_dc, 2)
    else:
        invoice['delivery_charge'] = 0.0

    for item in invoice.get('items', []):
        desc = item.get('description', '')
        bag_match = re.search(r'\(\s*(\d+(?:\.\d+)?)\s*Bags?\s*\)', desc, re.IGNORECASE)
        if bag_match and not item.get('bags'):
            item['bags'] = bag_match.group(1)
            item['description'] = re.sub(r'\s*\(\s*\d+(?:\.\d+)?\s*Bags?\s*\)', '', desc, flags=re.IGNORECASE).strip()

    buyers = db.get_all_buyers()
    matched_buyer = _match_buyer_profile(invoice, buyers)
    if matched_buyer:
        invoice['buyer_profile_id'] = matched_buyer.get('profile_id', '')
        invoice['buyer_name'] = matched_buyer.get('buyer_name') or invoice.get('buyer_name', '')
        invoice['buyer_gstin'] = matched_buyer.get('gstin') or invoice.get('buyer_gstin', '')
        invoice['buyer_details'] = matched_buyer.get('buyer_details', [])
    else:
        if not isinstance(invoice.get('buyer_details'), list):
            fallback_details = []
            if invoice.get('buyer_name'):
                fallback_details.extend(['Buyer :', invoice.get('buyer_name')])
            if invoice.get('buyer_gstin'):
                fallback_details.append(f"GSTIN - {invoice.get('buyer_gstin')}")
            invoice['buyer_details'] = fallback_details
    
    return jsonify(invoice)


@app.route('/api/invoice-file/<path:filename>')
def api_get_invoice_by_file(filename):
    """Load invoice data directly from a stored XLSX file."""
    safe_name = _safe_filename(filename, '.xlsx')
    file_bytes = get_cloud_storage().download_invoice_xlsx(safe_name)
    if not file_bytes:
        return jsonify({'error': 'Invoice file not found'}), 404

    try:
        invoice = _extract_invoice_data_from_xlsx_bytes(file_bytes, safe_name)
    except Exception as exc:
        app.logger.exception('Failed to parse XLSX %s: %s', safe_name, exc)
        return jsonify({'error': 'Failed to parse invoice file'}), 500

    buyers = get_sheets_db().get_all_buyers()
    matched_buyer = _match_buyer_profile(invoice, buyers)
    if matched_buyer:
        invoice['buyer_profile_id'] = matched_buyer.get('profile_id', '')
        invoice['buyer_name'] = matched_buyer.get('buyer_name') or invoice.get('buyer_name', '')
        invoice['buyer_gstin'] = matched_buyer.get('gstin') or invoice.get('buyer_gstin', '')
        invoice['buyer_details'] = matched_buyer.get('buyer_details', invoice.get('buyer_details', []))

    return jsonify(invoice)


@app.errorhandler(404)
def not_found(_error):
    """Render a friendly 404 page."""
    return render_template('error.html',
                          code=404,
                          title='Page Not Found',
                          message='The page you requested does not exist.'), 404


@app.errorhandler(429)
def too_many_requests(_error):
    """Render a friendly 429 page."""
    return render_template('error.html',
                          code=429,
                          title='Too Many Requests',
                          message='Please slow down and retry after a short pause.'), 429


@app.errorhandler(Exception)
def handle_unexpected_error(error):
    """Fallback error handler with logging for production debugging."""
    if isinstance(error, HTTPException):
        return error

    app.logger.exception('Unhandled error: %s', error)
    return render_template('error.html',
                          code=500,
                          title='Something Went Wrong',
                          message='An unexpected error occurred. Please try again.'), 500


# ===================== HEALTH CHECK =====================

@app.route('/health')
def health_check():
    """Health check endpoint for Cloud Run/App Engine."""
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})


@app.route('/favicon.ico')
def favicon():
    """Return empty favicon response to avoid repeated 404 noise in logs."""
    return ('', 204)


# ===================== MAIN =====================

if __name__ == '__main__':
    # For local development
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
