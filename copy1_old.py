import openpyxl
from copy import copy
from num2words import num2words

def copy_excel_with_formatting(source_filepath, destination_filepath, config):
    """
    Reads an Excel file, updates it with config data, and creates a copy with formatting preserved.

    Args:
        source_filepath (str): Path to the source Excel file.
        destination_filepath (str): Path to save the copied Excel file.
        config (dict): Dictionary containing the data to update in the Excel file.
    """
    try:
        source_wb = openpyxl.load_workbook(source_filepath)
    except FileNotFoundError:
        print(f"Error: Source file not found at {source_filepath}")
        return

    # Create a new workbook for the destination
    dest_wb = openpyxl.Workbook()
    # Remove the default sheet created with a new workbook
    if dest_wb.sheetnames: # Check if there's an active sheet to remove
        dest_wb.remove(dest_wb.active)

    for sheet_name in source_wb.sheetnames:
        source_sheet = source_wb[sheet_name]
        # Create a new sheet in the destination workbook with the same name
        dest_sheet = dest_wb.create_sheet(title=sheet_name)

        # Copy page setup (paper size, orientation, etc.)
        dest_sheet.page_setup = copy(source_sheet.page_setup)

        # Copy page margins
        dest_sheet.page_margins = copy(source_sheet.page_margins)

        # Copy cell values and styles (formatting)
        for row in source_sheet.iter_rows():
            for source_cell in row:
                # Create a new cell in the destination sheet and copy the value
                dest_cell = dest_sheet.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)

                # If the source cell has a style, copy it
                if source_cell.has_style:
                    dest_cell.font = copy(source_cell.font)
                    dest_cell.border = copy(source_cell.border)
                    dest_cell.fill = copy(source_cell.fill)
                    dest_cell.number_format = source_cell.number_format
                    dest_cell.protection = copy(source_cell.protection)
                    dest_cell.alignment = copy(source_cell.alignment)
        
        # Update Invoice Number and Date from config if provided
        # This should be AFTER the main cell copy loop to ensure these values are not overwritten.
        if config.get("invoice_number"):
            dest_sheet['E2'] = config.get("invoice_number")
        if config.get("invoice_date"):
            dest_sheet['H2'] = config.get("invoice_date") 
            # Ensure the cell H2 in your template is formatted to display dates correctly.

        # Update specific cells based on the config
        # Buyer Details (A8:A15)
        buyer_details = config.get("buyer_details", [])
        for i, detail in enumerate(buyer_details):
            dest_sheet[f'A{8+i}'] = detail

        # Mode of Transport (E10)
        dest_sheet['E10'] = config.get("mode_of_transport", "")

        # Item Details (Table A17:I17) - Assuming one item for now
        item = config.get("item_details", {})
        if item:
            dest_sheet['A18'] = item.get("description", "") # Description in A18 (assuming headers are in A17)
            dest_sheet['F18'] = item.get("quantity", 0)    # Quantity in F18
            dest_sheet['G18'] = item.get("rate", 0)        # Rate in G18
            # Calculate and set Amount (HSN/SAC * Rate) - Assuming HSN/SAC is in E18
            quantity = item.get("quantity", 0)
            rate = item.get("rate", 0)
            amount = quantity * rate
            dest_sheet['I18'] = amount                 # Amount in I18
            dest_sheet['I18'].number_format = '0.00'   # Ensure two decimal places
            
            # Update subtotal in I28 (assuming it's the sum of amounts, for one item it's just the amount)
            dest_sheet['I29'] = amount
            dest_sheet['I29'].number_format = '0.00'   # Ensure two decimal places


        # Tax Calculation (C29:I35)
        tax_type = config.get("tax_type", "IGST") # Default to IGST
        subtotal = dest_sheet['I29'].value or 0 # Get subtotal from I29

        igst_rate = 0
        cgst_rate = 0
        sgst_rate = 0

        if tax_type == "IGST":
            # IGST @ 5%
            igst_rate = 0.05
            dest_sheet['C30'] = "G.S.T SALES I.G.S.T @"
            dest_sheet['E30'] = "5.00%"
            dest_sheet['I30'] = subtotal * igst_rate
            dest_sheet['I30'].number_format = '0.00'
            dest_sheet['C31'] = "G.S.T SALES C.G.S.T @"
            dest_sheet['E31'] = "0.00%" # Corrected from 6.00% for IGST case
            dest_sheet['I31'] = 0.0
            dest_sheet['I31'].number_format = '0.00'
            dest_sheet['C32'] = "G.S.T SALES S.G.S.T @"
            dest_sheet['E32'] = "0.00%" # Corrected from 6.00% for IGST case
            dest_sheet['I32'] = 0.0
            dest_sheet['I32'].number_format = '0.00'
        elif tax_type == "CGST_SGST":
            # CGST 2.5% + SGST 2.5%
            cgst_rate = 0.025
            sgst_rate = 0.025
            dest_sheet['C30'] = "G.S.T SALES I.G.S.T @"
            dest_sheet['E30'] = "0.00%" # Corrected from 12.00% for CGST/SGST case
            dest_sheet['I30'] = 0.0
            dest_sheet['I30'].number_format = '0.00'
            dest_sheet['C31'] = "G.S.T SALES C.G.S.T @"
            dest_sheet['E31'] = "2.50%"
            dest_sheet['I31'] = subtotal * cgst_rate
            dest_sheet['I31'].number_format = '0.00'
            dest_sheet['C32'] = "G.S.T SALES S.G.S.T @"
            dest_sheet['E32'] = "2.50%"
            dest_sheet['I32'] = subtotal * sgst_rate
            dest_sheet['I32'].number_format = '0.00'

        igst_amount = dest_sheet['I30'].value or 0
        cgst_amount = dest_sheet['I31'].value or 0
        sgst_amount = dest_sheet['I32'].value or 0
        
        # Total Amount Calculation before round off
        total_before_round_off = subtotal + igst_amount + cgst_amount + sgst_amount
        
        # Round off calculation
        rounded_total = round(total_before_round_off)
        round_off_value = rounded_total - total_before_round_off
        
        dest_sheet['I34'] = round_off_value # Assuming I34 is for Round off
        dest_sheet['I34'].number_format = '0.00'   # Ensure two decimal places
        dest_sheet['I35'] = rounded_total   # Assuming I35 is for the final TOTAL
        dest_sheet['I35'].number_format = '0.00'   # Ensure two decimal places

        # Total Amount in Words: write next to the cell that contains 'AMOUNT'
        if rounded_total is not None:
            # Convert the integer part of the rounded total to words
            amount_in_words_str = num2words(int(rounded_total), lang='en_IN')
            # Remove hyphens and title case
            amount_in_words_str = amount_in_words_str.replace('-', ' ').replace(',', ' ').title()
            amount_in_words = amount_in_words_str + " Only"
        else:
            amount_in_words = "Zero Only"

        dest_sheet['A37'] = "AMOUNT : "+amount_in_words
        
        # Copy column dimensions (widths, hidden, etc.)
        for col_letter, source_dim in source_sheet.column_dimensions.items():
            dest_dim = dest_sheet.column_dimensions[col_letter]
            dest_dim.width = source_dim.width
            dest_dim.hidden = source_dim.hidden
            dest_dim.outline_level = source_dim.outline_level
            dest_dim.collapsed = source_dim.collapsed

        # Copy row dimensions (heights, hidden, etc.)
        for row_idx, source_dim in source_sheet.row_dimensions.items():
            dest_dim = dest_sheet.row_dimensions[row_idx]
            dest_dim.height = source_dim.height
            dest_dim.hidden = source_dim.hidden
            dest_dim.outline_level = source_dim.outline_level
            dest_dim.collapsed = source_dim.collapsed

        # Copy merged cells
        for merged_cell_range in source_sheet.merged_cells.ranges:
            dest_sheet.merge_cells(str(merged_cell_range))
            
    # Save the destination workbook
    try:
        dest_wb.save(destination_filepath)
        print(f"File copied successfully to {destination_filepath}")
    except Exception as e:
        print(f"Error saving destination file in copy1.py: {e}")
        raise # Re-raise the exception to be caught by app.py
    
if __name__ == '__main__':
    # Optional: manual test harness (disabled by default).
    pass